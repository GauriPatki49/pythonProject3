import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

def update_excel_data(filePath, searchName, ColName,new_Value):
    book = openpyxl.load_workbook("filePath")
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == "colName":
           Dict["col"] = i
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=1, column=j).value == "searchName":
               Dict["row"] = i
               
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = "new_Value"
    book.save(file_path)

driver=webdriver.Chrome()
driver.implicitly_wait(5)
file_path="C:\\Users\\swana\\Downloads\\download.xlsx"
fruit_name = "Apple"
newValue= "999"
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.find_element(By.ID, "downloadButton").click()
time.sleep(5)

#edit the Excel with updated value
update_excel_data(file_path,fruit_name, "Price", newValue)
file_input=driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator =(By.CSS_SELECTOR, "div[role='alert'] div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
assert actual_price == newValue
time.sleep(5)






