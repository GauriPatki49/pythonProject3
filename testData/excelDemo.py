import openpyxl
book = openpyxl.load_workbook("C:/Users/swana/Documents/PythonExcel.xlsx")
sheet = book.active
Dict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)
            
       
    
       