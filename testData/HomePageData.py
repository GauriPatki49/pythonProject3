import openpyxl
class HomePageData:
    
    Homepage_Data = [{"firstname": "Gauri", "lastname": "Mahurkar", "gender": "Female"},
                          {"firstname": "Swanand", "lastname": "Mahurkar", "gender": "Male"}]
    
    
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:/Users/swana/Documents/PythonExcel.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                var = Dict[sheet.cell(row=1, column=1).value]
               
            
            
            
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i,column=j).value
        return[Dict]